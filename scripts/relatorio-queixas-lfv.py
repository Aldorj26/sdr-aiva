# -*- coding: utf-8 -*-
# Gera o xlsx do relatório de queixas das lojas LFV a partir do JSON da mineração.
import json, io
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

dados = json.load(io.open('scripts/out-mineracao-queixas-lfv.json', encoding='utf-8'))

ROTULO = {
    'reprovacao_credito': 'Reprovação de crédito',
    'repasse_pagamento': 'Repasse / pagamento à loja',
    'ccb_sms_cliente_final': 'CCB / SMS do cliente final',
    'login_acesso_socio': 'Login / acesso do sócio',
    'cadastro_usuarios_vendedores': 'Cadastro de usuários (vendedores)',
    'plataforma_erro_tecnico': 'Erro técnico na plataforma',
    'boleto_parcela_cliente_final': 'Boleto / parcela do cliente final',
    'precificacao_taxa_mdr': 'Precificação / taxa / MDR',
    'material_divulgacao': 'Material de divulgação',
    'duvida_treinamento_operacao': 'Dúvida de treinamento / operação',
    'troca_conta_bancaria': 'Troca de conta bancária',
    'limite_valor_credito': 'Limite / valor de crédito',
    'cancelamento_estorno_troca': 'Cancelamento / estorno / troca',
    'equipe_desmotivada_parou_ofertar': 'Equipe desmotivada / parou de ofertar',
    'outros': 'Outros',
}
RESOLVEDOR = {
    'victoria': 'VictorIA', 'nei_manual': 'Nei (Enviar info)',
    'encaminhado_livechat_aiva': 'Encaminhado Live Chat AIVA',
    'encaminhado_outro_canal': 'Encaminhado outro canal',
    'nao_resolvido': 'NÃO resolvido', 'nao_da_pra_saber': 'Não dá pra saber',
}

probs = []          # linhas da aba Problemas
cat_lojas = defaultdict(set)
cat_ocorr = Counter()
resolvedor_cnt = Counter()
nao_resolvidos = []
nei_resolve = Counter()
reprov = []
lojas_ok = 0
erros = 0
por_loja = []

for lid, r in dados.items():
    nome = r.get('nome', '?'); tel = str(r.get('telefone', ''))
    if r.get('erro') or r.get('skip'):
        erros += 1
        continue
    lojas_ok += 1
    por_loja.append((nome, tel, r.get('n_msgs', 0), len(r.get('problemas', [])),
                     'sim' if r.get('sinais_reprovacao', {}).get('reclamou_reprovacao') else '',
                     'sim' if r.get('pediu_material_divulgacao') else '',
                     'sim' if r.get('elogiou_ou_vendeu_bem') else '',
                     r.get('resumo_loja', '')))
    for p in r.get('problemas', []):
        cat = p.get('categoria', 'outros')
        if cat not in ROTULO: cat = 'outros'
        res = p.get('resolvido_por', 'nao_da_pra_saber')
        probs.append((nome, tel, ROTULO[cat], p.get('descricao', ''), RESOLVEDOR.get(res, res),
                      'sim' if p.get('resposta_foi_boa') else 'não', p.get('evidencia', '')))
        cat_lojas[cat].add(nome); cat_ocorr[cat] += 1
        resolvedor_cnt[res] += 1
        if res == 'nao_resolvido':
            nao_resolvidos.append((nome, ROTULO[cat], p.get('descricao', ''), p.get('evidencia', '')))
        if res == 'nei_manual':
            nei_resolve[cat] += 1
    sr = r.get('sinais_reprovacao', {}) or {}
    if sr.get('reclamou_reprovacao') or sr.get('parou_ou_desanimou_de_ofertar'):
        reprov.append((nome, tel, sr.get('qtd_mencoes_reprovacao', 0),
                       'SIM' if sr.get('parou_ou_desanimou_de_ofertar') else 'não',
                       sr.get('evidencia', '')))

wb = Workbook()
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BASE = Font(name='Arial', size=10)

def aba(ws, headers, rows, larguras, telefone_cols=()):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical='center')
    for row in rows:
        ws.append(list(row))
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).font = BASE
            if c in telefone_cols:
                ws.cell(r, c).number_format = '@'
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

# ── Resumo ──
ws = wb.active; ws.title = 'Resumo'
tot_reprov_parou = sum(1 for x in reprov if x[3] == 'SIM')
linhas_resumo = [('Lojas analisadas (status Loja Finalizada e Vendendo)', lojas_ok),
                 ('Lojas sem conversa útil / erro de leitura', erros),
                 ('Problemas/queixas extraídos', len(probs)),
                 ('Lojas que reclamaram de reprovação de crédito', len(reprov)),
                 ('— destas, com sinal de desânimo/parou de ofertar', tot_reprov_parou),
                 ('Problemas SEM resolução identificada', len(nao_resolvidos)),
                 ('', ''),
                 ('RANKING DE CATEGORIAS', 'lojas afetadas | ocorrências')]
for cat, _ in cat_ocorr.most_common():
    linhas_resumo.append((ROTULO[cat], f'{len(cat_lojas[cat])} lojas | {cat_ocorr[cat]} ocorrências'))
linhas_resumo.append(('', ''))
linhas_resumo.append(('QUEM RESOLVEU', 'ocorrências'))
for res, n in resolvedor_cnt.most_common():
    linhas_resumo.append((RESOLVEDOR.get(res, res), n))
aba(ws, ['Indicador', 'Valor'], linhas_resumo, [58, 34])

aba(wb.create_sheet('Problemas'),
    ['Loja', 'Telefone', 'Categoria', 'Problema', 'Quem resolveu', 'Resposta boa?', 'Evidência (fala do lojista)'],
    sorted(probs, key=lambda x: x[2]), [30, 15, 30, 55, 22, 12, 50], telefone_cols=(2,))

aba(wb.create_sheet('Reprovação (barreira)'),
    ['Loja', 'Telefone', 'Menções a reprovação', 'Desanimou-parou de ofertar', 'Evidência'],
    sorted(reprov, key=lambda x: -(x[2] or 0)), [30, 15, 20, 24, 60], telefone_cols=(2,))

aba(wb.create_sheet('Não resolvidos'),
    ['Loja', 'Categoria', 'Problema', 'Evidência'],
    nao_resolvidos, [30, 30, 55, 50])

aba(wb.create_sheet('Nei via Enviar info'),
    ['Categoria', 'Ocorrências resolvidas pelo Nei manualmente'],
    [(ROTULO[c], n) for c, n in nei_resolve.most_common()], [40, 40])

aba(wb.create_sheet('Por loja'),
    ['Loja', 'Telefone', 'Msgs analisadas', 'Problemas', 'Reclamou reprovação', 'Pediu material', 'Vendendo bem-elogiou', 'Resumo'],
    sorted(por_loja, key=lambda x: -x[3]), [30, 15, 14, 11, 18, 13, 18, 70], telefone_cols=(2,))

out = 'docs/relatorio-queixas-lojas-lfv-2026-08-31.xlsx'
wb.save(out)
print('OK →', out)
print('lojas:', lojas_ok, '| problemas:', len(probs), '| reprovacao:', len(reprov), '| nao resolvidos:', len(nao_resolvidos))
