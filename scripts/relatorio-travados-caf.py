# -*- coding: utf-8 -*-
# Relatório xlsx dos leads travados no CAF (mineração 02/09).
import json, io, sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
dados = json.load(io.open('scripts/out-mineracao-travados-caf.json', encoding='utf-8'))

ROTULO = {
    'biometria_dificuldade_ou_medo': 'Biometria — dificuldade/receio',
    'link_ou_problema_tecnico': 'Link/problema técnico',
    'socio_ausente_sem_tempo': 'Sócio ausente / sem tempo',
    'esfriou_sem_motivo_declarado': 'Esfriou sem motivo declarado',
    'nunca_respondeu_cobrancas': 'Nunca respondeu às cobranças',
    'diz_que_vai_fazer_e_nao_faz': 'Diz que vai fazer e não faz',
    'aguardando_terceiro_ou_documento': 'Aguardando terceiro/documento',
    'desistiu_ou_mudou_de_ideia': 'Desistiu / mudou de ideia',
    'ja_diz_ter_concluido': '⚠️ Diz que JÁ CONCLUIU',
    'outro': 'Outro',
}
ACAO = {
    'ligar': '📞 Ligar', 'reenviar_link_com_passo_a_passo': '🔗 Reenviar link + passo a passo',
    'orientar_biometria': '🤳 Orientar biometria', 'esperar_data_combinada': '⏳ Esperar data combinada',
    'humano_negociar': '🧑‍💼 Humano negociar', 'considerar_descarte': '🗑️ Considerar descarte',
}

rows = []
cat_cnt = Counter(); acao_cnt = Counter(); vivos = 0
for lid, r in dados.items():
    if r.get('erro') or r.get('skip'): continue
    cat = r.get('categoria', 'outro')
    if cat not in ROTULO: cat = 'outro'
    acao = r.get('acao_recomendada', '')
    cat_cnt[cat] += 1; acao_cnt[acao] += 1
    if r.get('sinal_de_vida'): vivos += 1
    rows.append((r.get('nome',''), str(r.get('telefone','')), r.get('dias_na_etapa',''),
                 ROTULO[cat], 'sim' if r.get('respondeu_cobrancas') else 'NÃO',
                 'sim' if r.get('sinal_de_vida') else 'não',
                 r.get('motivo_resumo',''), ACAO.get(acao, acao), r.get('acao_detalhe',''),
                 r.get('ultima_fala_relevante','')))

rows.sort(key=lambda x: (x[7], -(x[2] if isinstance(x[2], int) else 0)))

wb = Workbook()
HDR_FILL = PatternFill('solid', fgColor='7C2D12'); HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BASE = Font(name='Arial', size=10)

def aba(ws, headers, data, larguras, tel_cols=()):
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = Alignment(vertical='center')
    for row in data: ws.append(list(row))
    for r in range(2, ws.max_row+1):
        for c in range(1, len(headers)+1):
            ws.cell(r, c).font = BASE
            if c in tel_cols: ws.cell(r, c).number_format = '@'
    for i, w in enumerate(larguras, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

ws = wb.active; ws.title = 'Resumo'
resumo = [('Leads travados analisados (3 cobranças esgotadas)', len(rows)),
          ('Com sinal de vida (ainda demonstram interesse)', vivos),
          ('', ''), ('POR MOTIVO', 'leads')]
for c, n in cat_cnt.most_common(): resumo.append((ROTULO[c], n))
resumo += [('', ''), ('POR AÇÃO RECOMENDADA', 'leads')]
for a, n in acao_cnt.most_common(): resumo.append((ACAO.get(a, a), n))
aba(ws, ['Indicador', 'Valor'], resumo, [52, 20])

aba(wb.create_sheet('Leads (lista de ação)'),
    ['Loja', 'Telefone', 'Dias na etapa', 'Motivo', 'Respondeu cobranças', 'Sinal de vida', 'Diagnóstico', 'Ação', 'Como agir', 'Última fala do lojista'],
    rows, [26, 15, 12, 30, 17, 12, 48, 26, 48, 45], tel_cols=(2,))

out = 'docs/relatorio-travados-caf-2026-09-02.xlsx'
wb.save(out)
print('OK ->', out)
print('total:', len(rows), '| vivos:', vivos)
for c, n in cat_cnt.most_common(): print(f'- {ROTULO[c]}: {n}')
print('acoes:')
for a, n in acao_cnt.most_common(): print(f'- {ACAO.get(a,a)}: {n}')
